__author__ = 'User'
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import smtplib
from email.mime.text import MIMEText
import base64
import imaplib
import email
import os
global _save_dir = os.path.dirname(os.path.realpath(__file__))

# gets a filename, retrieve the xlsx string data
def read_xlsx_file(filename1,filename2=""):
    rows = []
    if not filename1:
        wb = load_workbook(filename1)
        ws = wb.get_active_sheet()
        for row in ws.iter_rows(row_offset=1):
            rows.append(row)
        return rows,""
    return read_xlsx_file(filename1),read_xlsx_file(filename2)

# gets 2 filenames, returns a filename
def combine_xls(xl1, xl2):
    i=0
    data1 = read_xlsx_file(xl1)s
    data2 = read_xlsx_file(xl2)
    wb = Workbook()
    ws= wb.get_active_sheet()
    while st(i) in os.listdir(savedir):
        i++
    filename="/xl"+st(i)+".xlsx"
    openpyxl.workbook.save_workbook(_save_dir+filename)
    return filename
    
# gets type, sender, sender's password, recipients, message and a filename, returns a string
def send_to(t, sender, sender_pass, recipients, message, data):
    server = 'mail.server.com'
    user = sender
    password = sender_pass
    if t == "mail":
        try:
            smtplib.SMTP('localhost').sendmail(sender, recipients, message+'$'+base64.encode(data))
            return "success"
        except smtplib.SMTPException:
            return ""

# gets a recipent and a password and a t variable mentioning the way of sending, for now, it's limited to mail, or more accurately gmail
def receive(t, recipient, password):
    svdir = 'c:/downloads/'
    if type == "mail" and "gmail" in recipient:
        mail = imaplib.IMAP4_SSL('imap.gmail.com')
        mail.login(recipient, password)
        mail.list()  # Out: list of "folders" aka labels in gmail.
        mail.select("inbox")  # connect to inbox.
        result, data = mail.search(None, "ALL")

        ids = data[0]  # data is a list.
        id_list = ids.split()  # ids is a space separated string
        latest_email_id = id_list[-1]  # get the latest

        # fetch the email body (RFC822) for the given ID
        result, data = mail.fetch(latest_email_id, "(RFC822)")

        raw_email = data[0][1] # here's the body, which is raw text of the whole email
        #  including headers and alternate payloads

        typ, msgs = mail.search(None, '(SUBJECT "Detection")')
        msgs = msgs[0].split()

        for email_id in msgs:
            resp, data = mail.fetch(email_id, "(RFC822)")
            email_body = data[0][1]
            m = email.message_from_string(email_body)

            if m.get_content_maintype() != 'multipart':
                continue

            for part in m.walk():
                if part.get_content_maintype() == 'multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue

                filename = part.get_filename()
                if filename is not None:
                    sv_path = os.path.join(svdir, filename)
                    if not os.path.isfile(sv_path):
                        print sv_path
                        fp = open(sv_path, 'wb')
                        fp.write(part.get_payload(decode=True))
                        fp.close()

        return raw_email,filename
    return ""

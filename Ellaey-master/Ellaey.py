__author__ = 'Matan'
from openpyxl import load_workbook

from itertask import *
from bracha_on_office import *

"""
import socket

"""
"""
this is from the user perspective:

user 1 hr person -
on one place i have an excel of names family names emails and more data with a column for salary amount
"""

"""
# gets a file name and returns all data inscribed therein
def read_file_data(name):
    data = ""
    f = open(name,'r')
    for lines in f.readlines():
        data += lines
    f.close()
    return data


def read_xl(name):
    data = []
    f = load_workbook(name)
    wb = load_workbook(name, use_iterators=True)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    for row in ws.iter_rows(row_offset=1):
        data.append(row[0])


def read_files_data(names):
    dict_files_data = {}
    for i in len(names):
        dict_files_data[i] = (read_file_data(name=names[i]))
    return dict_files_data
"""

def main():
    _sf.append()
    TaskIterator()
    pass

if __name__ == "main":
    main()